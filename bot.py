from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes
from datetime import datetime, timedelta
import asyncio
import os
from openpyxl import Workbook, load_workbook

BOT_TOKEN = "7407040572:AAFuKe2eq74xCO024zy9OoEeSJ8Ojf34PpA"

# Penyimpanan data izin per grup
iizin_grup = {}       # chat_id -> {user_id: {jenis, start, limit}}
smoke_grup = {}       # chat_id -> [user_id1, user_id2]

# Waktu larangan merokok (dalam format HH:MM)
LARANGAN_SMOKE = [
    ("08:00", "08:20"),
    ("08:55", "09:25"),
    ("10:30", "11:00"),
    ("13:00", "13:30"),
    ("14:30", "14:45"),
    ("15:30", "16:00"),
    ("17:00", "17:30"),
    ("20:00", "20:15"),
    ("20:30", "21:00"),
    ("21:30", "22:00"),
    ("22:30", "23:00")
]

def dalam_jam_larangan():
    now = datetime.now().time()
    for mulai, akhir in LARANGAN_SMOKE:
        m = datetime.strptime(mulai, "%H:%M").time()
        a = datetime.strptime(akhir, "%H:%M").time()
        if m <= now <= a:
            return True
    return False

# Fungsi simpan log ke file Excel (xlsx)
def log_izin(timestamp, chat_id, user_id, nama, username, jenis, durasi, status):
    filename = "log_izin.xlsx"

    if os.path.exists(filename):
        wb = load_workbook(filename)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Timestamp", "Grup ID", "User ID", "Nama", "Username", "Jenis Izin", "Durasi", "Status"])

    sheet.append([
        timestamp,
        str(chat_id),
        str(user_id),
        nama,
        username or "-",
        jenis,
        str(durasi),
        status
    ])

    wb.save(filename)

# /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("🤖 Bot aktif! Gunakan /help untuk melihat perintah.")

# /help
async def help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📌 Perintah:\n"
        "/smoke - Ijin merokok (maks 2 orang/grup, 10 menit)\n"
        "/bab - Ijin buang air besar (15 menit)\n"
        "/toilet - Ijin ke toilet (5 menit)\n"
        "/1 - Kembali dari ijin\n"
        "/stat - Statistik pribadi hari ini\n"
        "/stat [user_id/@username] - Lihat stat orang lain\n"
        "/juarasmoke - Top 3 ijin merokok hari ini\n"
        "/juarabab - Top 3 ijin BAB hari ini\n"
        "/juaratoilet - Top 3 ijin toilet hari ini"
    )

# Fungsi pemberian ijin
async def beri_ijin(update, context, jenis, limit, emoji):
    user_id = update.effective_user.id
    nama = update.effective_user.first_name
    username = update.effective_user.username
    chat_id = update.effective_chat.id

    iizin_grup.setdefault(chat_id, {})
    smoke_grup.setdefault(chat_id, [])

    if user_id in iizin_grup[chat_id]:
        await update.message.reply_text("❗ Kamu masih dalam status ijin. Ketik /1 saat kembali.")
        return

    if jenis == "merokok":
        if dalam_jam_larangan():
            await update.message.reply_text(
                "❌ Kamu tidak diizinkan /Smoke sekarang.\n\n"
                "Dilarang mengajukan /Smoke di:\n"
                + "\n".join([f"• {m}-{a}" for m, a in LARANGAN_SMOKE])
            )
            return
        if len(smoke_grup[chat_id]) >= 1:
            await update.message.reply_text("🚫 Hanya 1 orang boleh merokok dalam satu grup saat ini.")
            return
        smoke_grup[chat_id].append(user_id)

    start_time = datetime.now()
    batas_waktu = start_time + timedelta(minutes=limit)

    iizin_grup[chat_id][user_id] = {
        "jenis": jenis,
        "start": start_time,
        "limit": limit
    }

    await update.message.reply_text(
        f"{emoji} {nama}, kamu ijin {jenis}.\n"
        f"Batas waktu: {limit} menit (sampai {batas_waktu.strftime('%H:%M:%S')}).\n"
        f"Ketik /1 saat kamu sudah kembali."
    )

    log_izin(
        timestamp=start_time.strftime('%Y-%m-%d %H:%M:%S'),
        chat_id=chat_id,
        user_id=user_id,
        nama=nama,
        username=username,
        jenis=jenis,
        durasi=0,
        status="Izin"
    )

    async def reminder():
        await asyncio.sleep((limit - 1) * 60)
        if user_id in iizin_grup.get(chat_id, {}):
            user_tag = f"@{username}" if username else nama
            await context.bot.send_message(
                chat_id=chat_id,
                text=f"⏰ {user_tag}, sisa waktu kamu hanya 1 menit lagi untuk {jenis}!"
            )

    asyncio.create_task(reminder())

# Handler ijin
async def smoke(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await beri_ijin(update, context, "merokok", 10, "🚬")

async def bab(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await beri_ijin(update, context, "buang air besar", 15, "🚽")

async def toilet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await beri_ijin(update, context, "toilet", 5, "🚻")

# Handler kembali
async def kembali(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    nama = update.effective_user.first_name
    username = update.effective_user.username
    chat_id = update.effective_chat.id

    if chat_id not in iizin_grup or user_id not in iizin_grup[chat_id]:
        await update.message.reply_text("❌ Kamu tidak dalam status ijin.")
        return

    info = iizin_grup[chat_id][user_id]
    start = info["start"]
    jenis = info["jenis"]
    limit = info["limit"]
    now = datetime.now()

    durasi = round((now - start).total_seconds() / 60, 1)
    terlambat = (now - start).total_seconds() > limit * 60

    if terlambat:
        telat = round(durasi - limit, 1)
        await update.message.reply_text(
            f"❌ {nama}, kamu terlambat!\n"
            f"Durasi: {durasi} menit\n"
            f"Keterlambatan: {telat} menit\n"
            f"💸 Denda: $10"
        )
        status = "Terlambat"
    else:
        await update.message.reply_text(f"✅ {nama}, kamu kembali tepat waktu. 🕒 Durasi: {durasi} menit.")
        status = "Tepat Waktu"

    log_izin(
        timestamp=now.strftime('%Y-%m-%d %H:%M:%S'),
        chat_id=chat_id,
        user_id=user_id,
        nama=nama,
        username=username,
        jenis=jenis,
        durasi=durasi,
        status=status
    )

    del iizin_grup[chat_id][user_id]
    if jenis == "merokok" and user_id in smoke_grup[chat_id]:
        smoke_grup[chat_id].remove(user_id)

# Handler statistik
async def stat(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    today = datetime.now().strftime('%Y-%m-%d')
    filename = "log_izin.xlsx"

    if not os.path.exists(filename):
        await update.message.reply_text("📄 Belum ada data izin untuk grup ini hari ini.")
        return

    target_user_id = update.effective_user.id
    target_nama = update.effective_user.first_name
    target_username = update.effective_user.username or "-"

    if context.args:
        arg = context.args[0]
        if arg.startswith("@"): 
            arg_username = arg[1:].lower()
        else:
            arg_username = None
            try:
                target_user_id = int(arg)
            except:
                await update.message.reply_text("❌ Format user ID tidak valid.")
                return
    else:
        arg_username = None

    wb = load_workbook(filename)
    sheet = wb.active

    nama_ditemukan = None
    uname_ditemukan = None
    jenis_durasi = {"merokok": 0.0, "buang air besar": 0.0, "toilet": 0.0}
    terlambat_count = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        timestamp, cid, uid, nama_log, uname_log, jenis, durasi, status = row

        if str(cid) != str(chat_id):
            continue
        if not str(timestamp).startswith(today):
            continue

        if arg_username:
            if str(uname_log).lower() != arg_username:
                continue
        else:
            if str(uid) != str(target_user_id):
                continue

        if not nama_ditemukan:
            nama_ditemukan = nama_log
            uname_ditemukan = uname_log

        if status != "Izin" and jenis in jenis_durasi:
            jenis_durasi[jenis] += float(durasi)
        if status == "Terlambat":
            terlambat_count += 1

    if nama_ditemukan is None:
        await update.message.reply_text("📄 Belum ada data izin untuk user ini hari ini.")
        return

    total = sum(jenis_durasi.values())
    denda = terlambat_count * 10
    teks = (
        f"📆 Statistik hari ini ({today}):\n"
        f"👤 {nama_ditemukan} (@{uname_ditemukan})\n"
        f"🚬 Merokok: {round(jenis_durasi['merokok'], 1)} menit\n"
        f"🚽 BAB: {round(jenis_durasi['buang air besar'], 1)} menit\n"
        f"🚻 Toilet: {round(jenis_durasi['toilet'], 1)} menit\n"
        f"⏱️ Total: {round(total, 1)} menit\n"
        f"❌ Terlambat: {terlambat_count} kali\n"
        f"💸 Total denda: ${denda}"
    )

    await update.message.reply_text(teks)

# Fungsi juara top 3
async def juara_handler(update: Update, context: ContextTypes.DEFAULT_TYPE, jenis_filter):
    chat_id = update.effective_chat.id
    today = datetime.now().strftime('%Y-%m-%d')
    filename = "log_izin.xlsx"

    if not os.path.exists(filename):
        await update.message.reply_text("Belum ada data ijin hari ini.")
        return

    wb = load_workbook(filename)
    sheet = wb.active
    durasi_per_user = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        timestamp, cid, uid, nama, uname, jenis, durasi, status = row
        if str(cid) != str(chat_id):
            continue
        if not str(timestamp).startswith(today):
            continue
        if status == "Izin":
            continue
        if jenis != jenis_filter:
            continue
        if uid not in durasi_per_user:
            durasi_per_user[uid] = {"nama": nama, "durasi": 0.0}
        durasi_per_user[uid]["durasi"] += float(durasi)

    hasil = sorted(durasi_per_user.items(), key=lambda x: x[1]['durasi'], reverse=True)[:3]

    if not hasil:
        await update.message.reply_text(f"Belum ada ijin {jenis_filter} hari ini.")
        return

    teks = f"🏆 Juara ijin {jenis_filter.upper()} hari ini:\n"
    for i, (uid, data) in enumerate(hasil, start=1):
        teks += f"{i}. {data['nama']} - {round(data['durasi'], 1)} menit\n"

    await update.message.reply_text(teks)

app = ApplicationBuilder().token(BOT_TOKEN).build()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("help", help))
app.add_handler(CommandHandler("smoke", smoke))
app.add_handler(CommandHandler("bab", bab))
app.add_handler(CommandHandler("toilet", toilet))
app.add_handler(CommandHandler("1", kembali))
app.add_handler(CommandHandler("stat", stat))
app.add_handler(CommandHandler("juarasmoke", lambda u, c: juara_handler(u, c, "merokok")))
app.add_handler(CommandHandler("juarabab", lambda u, c: juara_handler(u, c, "buang air besar")))
app.add_handler(CommandHandler("juaratoilet", lambda u, c: juara_handler(u, c, "toilet")))

print("🤖 Bot sedang berjalan...")
app.run_polling()